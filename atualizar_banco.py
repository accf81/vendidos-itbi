#!/usr/bin/env python3
"""
Pipeline Mensal ITBI SP
=======================
Baixa o arquivo XLSX do ano atual da Prefeitura de SP,
extrai os registros novos (residenciais) e insere no banco.

Uso:
  python3 atualizar_banco.py

O script:
  1. Baixa o XLSX do ano corrente
  2. Lê todas as abas (uma por mês)
  3. Filtra só residenciais (Descrição do uso IPTU = APARTAMENTO ou RESIDENCIAL)
  4. Descarta registros cujo SQL já está no banco (evita duplicatas)
  5. Aplica todas as normalizações (logradouro, cartório, numero)
  6. Insere os novos registros e recomprime o banco
"""

import sqlite3, gzip, shutil, re, os, time, sys, io
from datetime import datetime

try:
    import requests
except ImportError:
    print("Instalando requests..."); os.system("pip3 install requests -q")
    import requests

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl..."); os.system("pip3 install openpyxl -q")
    import openpyxl

# ─── Configuração ──────────────────────────────────────────────────
DB_FILE   = 'ITBI_SP_residencial.db.gz'
TMP_DB    = '/tmp/itbi_update.db'
BACKUP_FILE = f'ITBI_SP_residencial_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db.gz'

ANO_ATUAL = datetime.now().year
PAGINA_ITBI = 'https://prefeitura.sp.gov.br/web/fazenda/w/acesso_a_informacao/31501'
# O link direto do XLSX muda de endereço sem aviso (a Prefeitura publica um documento novo
# a cada atualização, não substitui o arquivo no mesmo lugar) — por isso o link é descoberto
# na página oficial a cada execução, em vez de fixo no código. Ver função obter_url_xlsx().

# Palavras-chave para identificar uso residencial (coluna "Descrição do uso (IPTU)")
# Os valores reais são textos longos: "APARTAMENTO EM CONDOMÍNIO (...)", "RESIDÊNCIA", etc.
USOS_RESIDENCIAIS_KW = ['APARTAMENTO', 'RESIDÊNCIA', 'RESIDENCIAL', 'FLAT RESIDENCIAL']

# Mapeamento de colunas do XLSX → campos do banco
# Baseado nas 28 colunas confirmadas do arquivo da Prefeitura (Jun/2026)
COLUNAS_XLSX = {
    'N° do Cadastro (SQL)':                'sql',
    'Nome do Logradouro':                  'logradouro',
    'Número':                              'numero',
    'Complemento':                         'complemento',
    'Bairro':                              'bairro',
    'Referência':                          'referencia',
    'Valor de Transação (declarado pelo contribuinte)': 'valor_transacao',
    'Data de Transação':                   'data_transacao',
    'Cartório de Registro':                'cartorio',
    'Matrícula do Imóvel':                 'matricula',
    'Área Construída (m2)':                'area_construida_m2',
    'Descrição do uso (IPTU)':             '_uso',  # filtro, não armazenado
}

# ─── Normalização (igual ao normalizar_banco.py) ────────────────────
NORMALIZE_MAP = [
    (r'^R\b','RUA'),(r'^AV\b','AVENIDA'),(r'^AL\b','ALAMEDA'),
    (r'^PC\b','PRACA'),(r'^PCA\b','PRACA'),(r'^TV\b','TRAVESSA'),
    (r'^TVP\b','TRAVESSA'),(r'^EST\b','ESTRADA'),(r'^ES\b','ESTRADA'),
    (r'^RD\b','RODOVIA'),(r'^RV\b','RODOVIA'),(r'^VL\b','VILA'),
    (r'^LG\b','LARGO'),(r'^VD\b','VIADUTO'),(r'^PQ\b','PARQUE'),
    (r'^LD\b','LADEIRA'),(r'^PS\b','PASSARELA'),(r'^VP\b','VIA PARQUE'),
    (r'\bNSRA\b','NOSSA SENHORA'),(r'\bNSA\b','NOSSA SENHORA'),
    (r'\bENG\b','ENGENHEIRO'),(r'\bENGA\b','ENGENHEIRA'),
    (r'\bDR\b','DOUTOR'),(r'\bDRA\b','DOUTORA'),
    (r'\bPROF\b','PROFESSOR'),(r'\bPROFA\b','PROFESSORA'),
    (r'\bPDE\b','PADRE'),(r'\bFRE\b','FREI'),(r'\bCEL\b','CORONEL'),
    (r'\bCAP\b','CAPITAO'),(r'\bTEN\b','TENENTE'),(r'\bBRIG\b','BRIGADEIRO'),
    (r'\bBR\b','BARAO'),(r'\bVIS\b','VISCONDE'),(r'\bCONS\b','CONSELHEIRO'),
    (r'\bMAL\b','MARECHAL'),(r'\bGEN\b','GENERAL'),(r'\bGOV\b','GOVERNADOR'),
    (r'\bPREF\b','PREFEITO'),(r'\bDEP\b','DEPUTADO'),(r'\bSEN\b','SENADOR'),
    (r'\bVER\b','VEREADOR'),(r'\bPRES\b','PRESIDENTE'),(r'\bMIN\b','MINISTRO'),
    (r'\bCOM\b','COMENDADOR'),(r'\bDES\b','DESEMBARGADOR'),(r'\bMAJ\b','MAJOR'),
    (r'\bDUQ\b','DUQUE'),(r'\bMARQ\b','MARQUES'),(r'\bCOND\b','CONDE'),
    (r'\bPRSA\b','PRINCESA'),(r'\bSTA\b','SANTA'),(r'\bSTO\b','SANTO'),
    (r'\bSTE\b','SANTO'),
]

def normalize_logradouro(s):
    if not s: return s
    s = str(s).upper().strip()
    s = re.sub(r'\s+', ' ', s)
    for pattern, replacement in NORMALIZE_MAP:
        s = re.sub(pattern, replacement, s)
    return s

def normalize_cartorio(s):
    if not s: return s
    m = re.search(r'(\d+)', str(s))
    return f"{m.group(1)} CRI" if m else str(s)

def parse_data(val):
    """Converte data do Excel para string YYYY-MM-DD HH:MM:SS"""
    if val is None: return None
    if hasattr(val, 'strftime'): return val.strftime('%Y-%m-%d %H:%M:%S')
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except: pass
    return s

def parse_float(val):
    if val is None: return None
    try: return float(str(val).replace('R$','').replace('.','').replace(',','.').strip())
    except: return None

def parse_str(val):
    if val is None: return None
    s = str(val).strip()
    return s.upper() if s else None

def ler_secret(nome):
    """Lê uma chave do arquivo local ~/.alex-os-secrets (formato NOME=valor, uma por linha).
    Nunca colar chave nenhuma direto no código — este arquivo é publicado num repositório
    público (accf81/vendidos-itbi)."""
    caminho = os.path.expanduser('~/.alex-os-secrets')
    if not os.path.exists(caminho):
        return None
    with open(caminho) as f:
        for linha in f:
            if linha.strip().startswith(f'{nome}='):
                return linha.strip().split('=', 1)[1]
    return None

def sincronizar_supabase(novos):
    """Envia os registros novos pra tabela vendas_itbi no Supabase — é o banco que a
    busca do site público (Pandora Data SP) consulta. Sem isso, o site fica com dado
    parado enquanto o arquivo local segue sendo atualizado (achado 15/07/2026, backlog
    15.1 do Alex OS). Precisa da SUPABASE_SERVICE_ROLE_KEY em ~/.alex-os-secrets — essa
    chave tem permissão de escrita (a chave pública do site só lê), pegar em
    supabase.com/dashboard/project/sobmjqounukzbplrmhkr/settings/api."""
    import json, urllib.request, urllib.error

    key = ler_secret('SUPABASE_SERVICE_ROLE_KEY')
    if not key:
        print(f"\n⚠️  Sincronização com o Supabase PULADA — falta SUPABASE_SERVICE_ROLE_KEY em ~/.alex-os-secrets.")
        print(f"    O banco local foi atualizado normalmente, mas a busca do site público")
        print(f"    (Pandora Data SP) vai continuar com dado antigo até isso ser configurado.")
        return

    url = 'https://sobmjqounukzbplrmhkr.supabase.co/rest/v1/vendas_itbi'
    headers = {'apikey': key, 'Authorization': f'Bearer {key}', 'Content-Type': 'application/json', 'Prefer': 'return=minimal'}
    campos = ['sql', 'data_transacao', 'logradouro', 'numero', 'complemento', 'bairro',
              'referencia', 'area_construida_m2', 'valor_transacao', 'valor_m2',
              'cartorio', 'matricula', 'ano', 'logradouro_norm']

    print(f"\n8. Sincronizando {len(novos):,} registros novos com o Supabase (site público)...")
    total = 0
    for i in range(0, len(novos), 1000):
        lote = novos[i:i+1000]
        payload = [dict(zip(campos, linha)) for linha in lote]
        req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), method='POST', headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                resp.read()
            total += len(lote)
        except urllib.error.HTTPError as e:
            print(f"   ✗ Erro no lote {i}-{i+len(lote)}: {e.code} {e.read().decode()[:300]}")
            print(f"    Rode a sincronização de novo depois, ou avise numa sessão DEV.")
            return
    print(f"   ✓ {total:,} registros sincronizados com o Supabase")

def obter_url_xlsx(ano, headers):
    """Busca na página oficial o link do XLSX do ano corrente — o endereço muda
    a cada republicação, não dá pra manter fixo no código (achado 15/07/2026:
    o link antigo estava parado numa versão de maio enquanto a Prefeitura já
    tinha publicado uma nova em julho, sem quebrar o download, só devolvendo
    dado velho)."""
    r = requests.get(PAGINA_ITBI, headers=headers, timeout=30)
    r.raise_for_status()
    m = re.search(rf'<strong>{ano}\s*\(<a href="([^"]+)"[^>]*>Excel/xlsx</a>', r.text)
    if not m:
        raise RuntimeError(f"Não achei o link do XLSX de {ano} na página oficial — o layout da página pode ter mudado.")
    url = m.group(1)
    if url.startswith('/'):
        url = 'https://prefeitura.sp.gov.br' + url
    return url

# ─── Main ──────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*55}")
    print(f"  PIPELINE MENSAL ITBI SP — {ANO_ATUAL}")
    print(f"{'='*55}\n")

    # 1. Backup e descomprimir banco atual
    print(f"1. Backup → {BACKUP_FILE}")
    shutil.copy2(DB_FILE, BACKUP_FILE)
    print(f"   ✓ {os.path.getsize(BACKUP_FILE)/1024/1024:.1f} MB")

    print(f"\n2. Descomprimindo banco...")
    with gzip.open(DB_FILE, 'rb') as f: data = f.read()
    with open(TMP_DB, 'wb') as f: f.write(data)
    print(f"   ✓ {len(data)/1024/1024:.1f} MB")

    conn = sqlite3.connect(TMP_DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM vendas"); total_antes = cur.fetchone()[0]
    print(f"   Registros antes: {total_antes:,}")

    # 2. Carregar chaves existentes para deduplicação rápida
    # Chave = (sql, numero, complemento) — um mesmo código de SQL (cadastro/lote) pode
    # cobrir várias unidades vendidas juntas (prédio inteiro, por exemplo); usar só o
    # sql como chave descartaria erroneamente as unidades além da primeira.
    print(f"\n3. Carregando registros existentes...")
    cur.execute("SELECT sql, numero, complemento FROM vendas WHERE sql IS NOT NULL")
    chaves_existentes = {(r[0], r[1], r[2]) for r in cur.fetchall()}
    print(f"   ✓ {len(chaves_existentes):,} registros carregados")

    # 3. Descobrir e baixar XLSX
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'}
    print(f"\n4. Descobrindo o link do XLSX {ANO_ATUAL} na página oficial...")
    try:
        url_xlsx = obter_url_xlsx(ANO_ATUAL, headers)
        print(f"   ✓ {url_xlsx}")
    except Exception as e:
        print(f"   ✗ Erro ao achar o link: {e}")
        conn.close(); os.remove(TMP_DB)
        sys.exit(1)

    print(f"\n   Baixando XLSX {ANO_ATUAL} da Prefeitura...")
    try:
        r = requests.get(url_xlsx, headers=headers, timeout=120, allow_redirects=True)
        r.raise_for_status()
        print(f"   ✓ {len(r.content)/1024/1024:.1f} MB baixados")
    except Exception as e:
        print(f"   ✗ Erro ao baixar: {e}")
        conn.close(); os.remove(TMP_DB)
        sys.exit(1)

    # 4. Parsear XLSX
    print(f"\n5. Lendo XLSX...")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    print(f"   Abas encontradas: {wb.sheetnames}")

    novos = []
    usos_encontrados = set()
    total_lido = total_residencial = total_duplicata = total_invalido = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            print(f"   ⚠ Aba '{sheet_name}' vazia, pulando.")
            continue

        # Mapear posição das colunas pelo nome
        col_idx = {}
        for i, nome in enumerate(header):
            if nome is None: continue
            nome_str = str(nome).strip()
            if nome_str in COLUNAS_XLSX:
                col_idx[COLUNAS_XLSX[nome_str]] = i

        # Verificar colunas obrigatórias
        obrigatorias = ['sql', 'logradouro', 'valor_transacao', 'data_transacao', '_uso']
        faltando = [c for c in obrigatorias if c not in col_idx]
        if faltando:
            print(f"   ⚠ Aba '{sheet_name}': colunas não encontradas: {faltando}")
            print(f"     Cabeçalho encontrado: {[str(h) for h in header if h]}")
            continue

        # Processar linhas
        for row in rows_iter:
            if not any(row): continue
            total_lido += 1

            uso = parse_str(row[col_idx['_uso']]) or ''
            usos_encontrados.add(uso)

            # Filtro residencial — verifica se alguma palavra-chave está no texto do uso
            if not any(kw in uso for kw in USOS_RESIDENCIAIS_KW):
                continue
            total_residencial += 1

            sql_val = parse_str(row[col_idx['sql']])
            if not sql_val:
                total_invalido += 1; continue

            # Extrair campos
            def get(field, parser=parse_str):
                idx = col_idx.get(field)
                return parser(row[idx]) if idx is not None else None

            logradouro    = get('logradouro')
            numero        = get('numero')
            complemento   = get('complemento')

            numero_norm = 'S/N' if str(numero or '').strip() == '99999' else str(numero or '').strip()

            # Deduplicação — por (sql, numero, complemento), não só sql (ver nota acima)
            chave = (sql_val, numero_norm, complemento)
            if chave in chaves_existentes:
                total_duplicata += 1; continue

            bairro        = get('bairro')
            referencia    = get('referencia')
            valor         = get('valor_transacao', parse_float)
            data_str      = get('data_transacao', parse_data)
            cartorio      = get('cartorio')
            matricula     = get('matricula')
            area          = get('area_construida_m2', parse_float)

            # Validação mínima
            if valor is None or valor < 100: total_invalido += 1; continue

            # Normalizações
            logradouro_norm = normalize_logradouro(logradouro)
            cartorio_norm   = normalize_cartorio(cartorio)
            valor_m2        = round(valor / area, 2) if area and area > 0 else None

            # Ano a partir da data
            try:
                ano = int(data_str[:4]) if data_str else ANO_ATUAL
            except:
                ano = ANO_ATUAL

            novos.append((
                sql_val, data_str, logradouro, numero_norm, complemento,
                bairro, referencia, area, valor, valor_m2,
                cartorio_norm, matricula, ano, logradouro_norm
            ))
            chaves_existentes.add(chave)

    print(f"\n   Estatísticas de leitura:")
    print(f"   Total linhas lidas:    {total_lido:,}")
    print(f"   Residenciais:          {total_residencial:,}")
    print(f"   Duplicatas (já no DB): {total_duplicata:,}")
    print(f"   Inválidos:             {total_invalido:,}")
    print(f"   → Novos a inserir:     {len(novos):,}")
    print(f"\n   Tipos de uso encontrados no XLSX: {sorted(usos_encontrados)}")

    if not novos:
        print("\n✅ Nenhum registro novo. Banco já está atualizado.")
        conn.close(); os.remove(TMP_DB); os.remove(BACKUP_FILE)
        return

    # 5. Inserir novos registros
    print(f"\n6. Inserindo {len(novos):,} registros...")
    cur.executemany("""
        INSERT OR IGNORE INTO vendas
          (sql, data_transacao, logradouro, numero, complemento,
           bairro, referencia, area_construida_m2, valor_transacao, valor_m2,
           cartorio, matricula, ano, logradouro_norm)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, novos)
    conn.commit()

    cur.execute("SELECT COUNT(*) FROM vendas"); total_depois = cur.fetchone()[0]
    inseridos = total_depois - total_antes
    print(f"   ✓ {inseridos:,} registros inseridos (total: {total_depois:,})")

    conn.close()

    sincronizar_supabase(novos)

    # 6. Recomprimir
    print(f"\n7. Recomprimindo banco...")
    t0 = time.time()
    with open(TMP_DB, 'rb') as f_in:
        with gzip.open(DB_FILE, 'wb', compresslevel=6) as f_out:
            shutil.copyfileobj(f_in, f_out)
    print(f"   ✓ {os.path.getsize(DB_FILE)/1024/1024:.1f} MB em {time.time()-t0:.1f}s")
    os.remove(TMP_DB)

    print(f"\n{'─'*40}")
    print(f"  Registros antes:  {total_antes:,}")
    print(f"  Novos inseridos:  {inseridos:,}")
    print(f"  Total final:      {total_depois:,}")
    print(f"{'─'*40}")
    print(f"\n✅ CONCLUÍDO!")
    print(f"   Backup mantido: {BACKUP_FILE}")
    print(f"\n   Próximo passo: duplo clique em deploy_github.command\n")

if __name__ == '__main__':
    main()
