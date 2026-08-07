#!/usr/bin/env python3
"""
Leitor v2 das planilhas do ITBI — base definitiva (TODOS os tipos de imóvel, 28 colunas).
=========================================================================================
Gera um banco NOVO a partir das planilhas originais da Prefeitura. NÃO toca em nada em
produção: escreve num arquivo à parte e grava um BOLETIM em Markdown do que leu e do que
NÃO leu (lição L-21 — importação que "pula o que não entende" apaga dado em silêncio).

O que muda em relação ao reconstruir_banco.py (que continua intacto, é a volta atrás):
  1. Sem filtro de uso — terreno, loja, galpão, escola, hotel: tudo entra. O tipo virou
     coluna (`descricao_uso_iptu`), não filtro. `residencial` fica como marca 0/1 derivada.
  2. Guarda as 28 colunas da planilha, não 13.
  3. Cada registro carrega a origem: `ano_planilha` (arquivo lido) e `aba_origem` (nome da
     aba). O `ano` do registro continua vindo da DATA DA TRANSAÇÃO — uma venda de dezembro
     declarada em janeiro cai num balde e vem de outro arquivo; sem a origem não dá pra
     explicar diferença de contagem.
  4. Deduplicação dentro do arquivo SQLite (índice único), não em dicionário na memória —
     são 2,7 milhões de linhas. Chave igual à de hoje: (sql, numero, complemento, data, valor).

O problema que este programa existe pra resolver: o layout varia DE ABA PRA ABA dentro do
mesmo arquivo e o rótulo da coluna MENTE.
  - 2019 a 2022 (todas as abas), 2 abas de 2023 e 3 de 2024 trazem `ACC (IPTU)` DUAS vezes
    (posições 26 e 27) e não trazem `Descrição do padrão (IPTU)`. O dado está lá: a posição
    26 é a descrição do padrão (RESIDENCIAL VERTICAL / HORIZONTAL / TERRENO) com o rótulo
    errado, e a 27 é o ano. Casar por nome exato grava "RESIDENCIAL VERTICAL" num campo de
    ano, sem erro nenhum.
  - 2026 traz `Descrição do pardão (IPTU)` — erro de digitação da Prefeitura.
  - JAN-2024 e OUT-2024 NÃO TÊM linha de cabeçalho: a primeira linha já é dado. São as
    ~20 mil vendas de 2024 que nunca entraram no banco.
Por isso: layout decidido POR ABA, rótulo casado de forma tolerante (sem acento, sem caixa,
com lista de grafias erradas) e, quando o rótulo é duplicado ou suspeito, o CONTEÚDO desempata.

Regras de tratamento preservadas (ITBI_REGRAS_E_HISTORICO.md §2), reusando as funções que
já existem — nada reescrito à mão:
  - valor respeitando o tipo da célula (bug x100)            [ab.parse_float]
  - tira o ".0" de campo de código vindo como número         [ab.tirar_decimal_excel/parse_str]
  - 99999 -> S/N                                             [ab.normalize_numero]
  - normalização de logradouro + as 34 correções             [ab.NORMALIZE_MAP + ab.SPECIFIC_CORRECTIONS]
    (+ 8 abreviações novas, ver EXTRA_NORMALIZE_MAP)
  - cartório em "N CRI"                                      [ab.normalize_cartorio]
  - valor_m2 calculado                                       [aqui]
  - logradouro_fmt pra exibição                              [nb.title_case_logradouro]
  - ano derivado da data                                     [aqui]

Uso:
  python3 reconstruir_banco_v2.py --anos 2019 --saida /tmp/itbi_2019.db
  python3 reconstruir_banco_v2.py --anos 2019-2026 --saida /tmp/itbi.db --boletim /tmp/bol.md
  python3 reconstruir_banco_v2.py --arquivos /caminho/2019_teste.xlsx --saida /tmp/t.db

Sai com código 1 (erro) quando alguma aba TEM linhas e NENHUMA foi entendida — é a
assinatura exata do JAN-2024. Aba vazia na origem (mês sem lançamento) e aba que só não
trouxe novidade NÃO são erro.
"""
import argparse
import itertools
import os
import re
import sqlite3
import statistics
import sys
import time
import unicodedata
from datetime import datetime

import openpyxl

import atualizar_banco as ab
import normalizar_banco as nb

PLAN_DIR = "/Users/accf81/Documents/IA/Claude/Projects/Dados ITBI/planilhas_originais"

VALOR_MINIMO = 100        # abaixo disso não é preço de imóvel (regra do banco atual)
# Decisão do Alex: NÃO filtrar valor absurdo — guardar tudo e poder ver tudo. Mas também não
# pode passar despercebido: acima deste limiar o boletim conta e lista os dez maiores.
LIMIAR_VALOR_ABSURDO = 1_000_000_000     # R$ 1 bilhão
# Mês abaixo desta fração da mediana do próprio ano vira aviso no boletim (detector de buraco —
# é o que teria pegado JAN/OUT-2024 há um ano).
FRACAO_MES_SUSPEITO = 0.40
AMOSTRA_LINHAS = 400      # linhas guardadas pra decidir o layout pelo conteúdo
MIN_ROTULOS_CABECALHO = 10  # abaixo disso, a primeira linha não é cabeçalho: já é dado

# ─── As 28 colunas da planilha, na ordem oficial ────────────────────────────────
# (campo no banco, rótulo oficial da Prefeitura, como ler o valor)
#   txt    = texto em MAIÚSCULAS, sem o ".0" do Excel
#   num    = número (respeitando o tipo da célula — regra 2.9)
#   data   = data
#   numero = número do imóvel (99999 -> S/N)
CANONICAS = [
    ('sql',                                 'N° do Cadastro (SQL)',                             'txt'),
    ('logradouro',                          'Nome do Logradouro',                               'txt'),
    ('numero',                              'Número',                                           'numero'),
    ('complemento',                         'Complemento',                                      'txt'),
    ('bairro',                              'Bairro',                                           'txt'),
    ('referencia',                          'Referência',                                       'txt'),
    ('cep',                                 'CEP',                                              'txt'),
    ('natureza_transacao',                  'Natureza de Transação',                            'txt'),
    ('valor_transacao',                     'Valor de Transação (declarado pelo contribuinte)', 'num'),
    ('data_transacao',                      'Data de Transação',                                'data'),
    ('valor_venal_referencia',              'Valor Venal de Referência',                        'num'),
    ('proporcao_transmitida',               'Proporção Transmitida (%)',                        'num'),
    ('valor_venal_referencia_proporcional', 'Valor Venal de Referência (proporcional)',         'num'),
    ('base_calculo_adotada',                'Base de Cálculo adotada',                          'num'),
    ('tipo_financiamento',                  'Tipo de Financiamento',                            'txt'),
    ('valor_financiado',                    'Valor Financiado',                                 'num'),
    ('cartorio',                            'Cartório de Registro',                             'txt'),
    ('matricula',                           'Matrícula do Imóvel',                              'txt'),
    ('situacao_sql',                        'Situação do SQL',                                  'txt'),
    ('area_terreno_m2',                     'Área do Terreno (m2)',                             'num'),
    ('testada_m',                           'Testada (m)',                                      'num'),
    ('fracao_ideal',                        'Fração Ideal',                                     'num'),
    ('area_construida_m2',                  'Área Construída (m2)',                             'num'),
    ('uso_iptu',                            'Uso (IPTU)',                                       'txt'),
    ('descricao_uso_iptu',                  'Descrição do uso (IPTU)',                          'txt'),
    ('padrao_iptu',                         'Padrão (IPTU)',                                    'txt'),
    ('descricao_padrao_iptu',               'Descrição do padrão (IPTU)',                       'txt'),
    ('acc_iptu',                            'ACC (IPTU)',                                       'txt'),
]
CAMPOS_ORIGEM = [c[0] for c in CANONICAS]
TIPO_CAMPO = {c[0]: c[2] for c in CANONICAS}
POSICAO_CANONICA = {c[0]: i for i, c in enumerate(CANONICAS)}

# Abas que não são de dados — tratadas pelo nome, não são falha.
ABAS_NAO_DADOS = ['LEGENDA', 'EXPLICACOES', 'TABELA DE USOS', 'TABELA DE PADROES']

# Campos sem os quais a linha não foi entendida (o layout não pegou).
CAMPOS_ESSENCIAIS = ['sql', 'valor_transacao']

# ─── 8 abreviações que a normalização de hoje não expande (476 ruas afetadas) ───
# CON fica de fora de propósito: é ambíguo (Conselheiro ou Cônego), decisão caso a caso.
EXTRA_NORMALIZE_MAP = [
    (r'\bGAL\b',  'GENERAL'),
    (r'\bSOLD\b', 'SOLDADO'),
    (r'\bMONS\b', 'MONSENHOR'),
    (r'\bVISC\b', 'VISCONDE'),
    (r'\bTTE\b',  'TENENTE'),
    (r'\bSARG\b', 'SARGENTO'),
    (r'\bALM\b',  'ALMIRANTE'),
    (r'\bARQ\b',  'ARQUITETO'),
]
NORMALIZE_MAP_V2 = list(ab.NORMALIZE_MAP) + EXTRA_NORMALIZE_MAP


def normalize_logradouro_v2(s):
    """Igual ao ab.normalize_logradouro, com as 8 abreviações novas no fim do mapa.
    O script de hoje (atualizar_banco.py) não é alterado — este é o mapa da base nova."""
    if not s:
        return s
    s = str(s).upper().strip()
    s = re.sub(r'\s+', ' ', s)
    for pattern, replacement in NORMALIZE_MAP_V2:
        s = re.sub(pattern, replacement, s)
    return ab.SPECIFIC_CORRECTIONS.get(s, s)


# ─── Casamento tolerante de rótulo ──────────────────────────────────────────────
def chave_rotulo(s):
    """'Descrição do pardão (IPTU)' -> 'DESCRICAO DO PARDAO IPTU'.
    Tira acento, caixa, pontuação e espaço extra — o rótulo da Prefeitura varia."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^A-Za-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip().upper()


# Grafias erradas conhecidas da Prefeitura (chave já normalizada)
GRAFIAS_ERRADAS = {
    'DESCRICAO DO PARDAO IPTU': 'DESCRICAO DO PADRAO IPTU',
    'DESCICAO DO PADRAO IPTU': 'DESCRICAO DO PADRAO IPTU',
    'DESCRICAO DO USO IPTU ': 'DESCRICAO DO USO IPTU',
}
POR_ROTULO = {chave_rotulo(rot): campo for campo, rot, _ in CANONICAS}


def campo_do_rotulo(rotulo):
    k = chave_rotulo(rotulo)
    if not k:
        return None
    k = GRAFIAS_ERRADAS.get(k, k)
    k = k.replace('PARDAO', 'PADRAO')   # rede extra pro erro de digitação
    return POR_ROTULO.get(k)


# ─── Desempate pelo conteúdo (descrição do padrão × ACC) ────────────────────────
RE_SO_NUMERO = re.compile(r'^\d+(\.0+)?$')
PALAVRAS_PADRAO = ('VERTICAL', 'HORIZONTAL', 'TERRENO')


def _assinatura(valores):
    """Diz quantas células parecem descrição de padrão e quantas parecem ano (ACC)."""
    n_padrao = n_ano = 0
    for v in valores:
        if v is None:
            continue
        s = str(v).strip().upper()
        if not s:
            continue
        if RE_SO_NUMERO.fullmatch(s):
            try:
                n = int(float(s))
            except ValueError:
                continue
            if 1500 <= n <= 2100:
                n_ano += 1
        elif any(p in s for p in PALAVRAS_PADRAO):
            n_padrao += 1
    return n_padrao, n_ano


def desempatar_padrao_acc(amostra, ncols):
    """Descobre pelo CONTEÚDO qual coluna é a descrição do padrão e qual é o ACC.
    Descrição do padrão é texto terminando em VERTICAL/HORIZONTAL (ou TERRENO);
    ACC é ano de 4 dígitos. Olha só as últimas 6 colunas — é onde as duas vivem."""
    ini = max(0, ncols - 6)
    melhor_padrao = (0, None)
    melhor_ano = (0, None)
    for i in range(ini, ncols):
        col = [linha[i] if i < len(linha) else None for linha in amostra]
        n_padrao, n_ano = _assinatura(col)
        if n_padrao > melhor_padrao[0]:
            melhor_padrao = (n_padrao, i)
        if n_ano > melhor_ano[0]:
            melhor_ano = (n_ano, i)
    col_padrao = melhor_padrao[1]
    col_acc = melhor_ano[1]
    if col_acc is not None and col_acc == col_padrao:
        col_acc = None
    return col_padrao, col_acc


# ─── Resolução do layout de UMA aba ─────────────────────────────────────────────
class Layout:
    def __init__(self):
        self.mapa = {}
        self.tem_cabecalho = True
        self.descricao = ''
        self.observacoes = []
        self.ok = False
        self.motivo = ''


def resolver_layout(cabecalho, amostra, ncols):
    """Decide, PARA ESTA ABA, qual posição é cada campo. Nunca em silêncio: tudo o que
    foi inferido ou corrigido pelo conteúdo entra em `observacoes` e vai pro boletim."""
    lay = Layout()
    rotulos = list(cabecalho) if cabecalho else []
    achados = 0
    duplicados = []
    for i, rot in enumerate(rotulos):
        campo = campo_do_rotulo(rot)
        if not campo:
            continue
        achados += 1
        if campo in lay.mapa:
            duplicados.append((campo, i))
        else:
            lay.mapa[campo] = i

    lay.tem_cabecalho = achados >= MIN_ROTULOS_CABECALHO

    if lay.tem_cabecalho:
        lay.descricao = (f'cabeçalho reconhecido ({len(lay.mapa)} de {len(CANONICAS)} rótulos'
                         + (f', {len(duplicados)} repetido(s)' if duplicados else '') + ')')
        if duplicados:
            nomes = ', '.join(sorted({d[0] for d in duplicados}))
            lay.observacoes.append(f'rótulo repetido no cabeçalho: {nomes}')
        faltando = [c for c in CAMPOS_ORIGEM if c not in lay.mapa]
        if faltando:
            lay.observacoes.append('rótulo ausente: ' + ', '.join(faltando))
    else:
        # Sem cabeçalho: a primeira linha já é dado. Assume o layout canônico de 28
        # colunas POR POSIÇÃO e confere pelo conteúdo antes de aceitar.
        if ncols < len(CANONICAS):
            lay.motivo = (f'sem cabeçalho reconhecível e só {ncols} colunas '
                          f'(o layout canônico tem {len(CANONICAS)})')
            return lay
        lay.mapa = dict(POSICAO_CANONICA)
        lay.descricao = f'SEM cabeçalho — layout inferido por posição (28 colunas canônicas, aba com {ncols})'
        i_data = POSICAO_CANONICA['data_transacao']
        i_sql = POSICAO_CANONICA['sql']
        datas_ok = sum(1 for l in amostra[:100]
                       if l[i_data] is not None and hasattr(l[i_data], 'strftime'))
        sqls_ok = sum(1 for l in amostra[:100] if l[i_sql] not in (None, ''))
        base = min(100, len(amostra))
        if base == 0 or datas_ok < base * 0.8 or sqls_ok < base * 0.8:
            lay.motivo = (f'sem cabeçalho e o conteúdo não confere com o layout canônico '
                          f'(data válida em {datas_ok}/{base}, cadastro em {sqls_ok}/{base})')
            return lay
        lay.observacoes.append(
            f'conferido pelo conteúdo: data válida em {datas_ok}/{base} linhas, cadastro em {sqls_ok}/{base}')

    # Desempate pelo conteúdo — vale sempre, inclusive pra desmentir rótulo que existe.
    col_padrao, col_acc = desempatar_padrao_acc(amostra, ncols)
    for campo, achado in (('descricao_padrao_iptu', col_padrao), ('acc_iptu', col_acc)):
        if achado is None:
            continue
        antes = lay.mapa.get(campo)
        if antes != achado:
            lay.mapa[campo] = achado
            if antes is None:
                lay.observacoes.append(f'{campo}: sem rótulo, achado pelo conteúdo na posição {achado}')
            else:
                lay.observacoes.append(
                    f'{campo}: rótulo apontava a posição {antes}, o conteúdo diz {achado} — vale o conteúdo')
    # Se o ACC não apareceu na amostra (aba só de terreno, ACC vazio), cai pra posição
    # seguinte à da descrição do padrão — e avisa.
    if col_acc is None and col_padrao is not None and col_padrao + 1 < ncols:
        if lay.mapa.get('acc_iptu') in (None, col_padrao):
            lay.mapa['acc_iptu'] = col_padrao + 1
            lay.observacoes.append(
                f'acc_iptu: nenhum ano na amostra; assumida a posição {col_padrao + 1} (logo após a descrição do padrão)')

    faltam = [c for c in CAMPOS_ESSENCIAIS + ['data_transacao'] if c not in lay.mapa]
    if faltam:
        lay.motivo = 'faltou coluna essencial: ' + ', '.join(faltam)
        return lay
    lay.ok = True
    return lay


# ─── Leitura ────────────────────────────────────────────────────────────────────
class ContaAba:
    def __init__(self, nome):
        self.nome = nome
        self.lidas = 0
        self.entendidas = 0
        self.novas = 0
        self.repetidas = 0
        self.residenciais_lidas = 0
        self.abaixo_do_piso = 0    # entram, mas sem valor por m² (plano §3.3(h))
        self.sem_cadastro = 0      # entram assim mesmo
        self.descartes = {}
        self.layout = ''
        self.observacoes = []
        self.lida = True
        self.motivo = ''
        self.eh_dados = True

    def descarte(self, motivo):
        self.descartes[motivo] = self.descartes.get(motivo, 0) + 1

    def situacao(self):
        """Como esta aba terminou — é o que o boletim precisa dizer com todas as letras."""
        if not self.eh_dados:
            return 'apoio'
        if not self.lida:
            return 'nao_lida'          # layout não reconhecido — as linhas ficaram de fora
        if self.lidas == 0:
            return 'vazia'             # mês sem lançamento — não é falha
        if self.entendidas == 0:
            return 'nada_entendido'    # tem linha e não aproveitou nenhuma — é falha (veto)
        return 'lida'


def eh_aba_de_dados(nome):
    k = chave_rotulo(nome)
    return not any(k.startswith(p) for p in ABAS_NAO_DADOS)


def _cel(linha, idx):
    return linha[idx] if idx is not None and idx < len(linha) else None


# A gravação é por ABA, não por lote menor: a maior aba da série tem 21 mil linhas, o que cabe
# folgado na memória. Uma tentativa de gravar de 4 mil em 4 mil foi começada e abandonada no meio
# (a assinatura de ler_aba chegou a receber a conexão, o corpo nunca usou) — desfeita em 06/08 por
# quebrar a chamada. Se um dia a memória apertar de novo, o lugar certo de mexer é aqui, e a causa
# do aperto de 06/08 foi outra: cache do SQLite grande demais, já reduzido em criar_banco().


def ler_aba(ws, nome_aba, ano_planilha, conta):
    """Lê a aba e vai gravando em lotes. Nunca pula aba em silêncio: o que não foi lido
    fica registrado em `conta` e vai pro boletim."""
    it = ws.iter_rows(values_only=True)
    primeira = next(it, None)
    if primeira is None:
        conta.layout = '—'
        conta.motivo = 'aba vazia na origem (nenhuma linha)'
        return []

    ncols = len(primeira)
    # Guarda uma amostra pra decidir o layout pelo conteúdo — uma passada só.
    amostra = []
    consumidas = []
    for linha in it:
        consumidas.append(linha)
        if any(linha):
            amostra.append(linha)
        if len(amostra) >= AMOSTRA_LINHAS:
            break
    # Se a primeira linha for dado (aba sem cabeçalho), ela também entra na amostra.
    parece_cabecalho = sum(1 for r in primeira if campo_do_rotulo(r)) >= MIN_ROTULOS_CABECALHO
    if not parece_cabecalho and any(primeira):
        amostra.insert(0, primeira)

    lay = resolver_layout(primeira, amostra, ncols)
    conta.layout = lay.descricao or lay.motivo
    conta.observacoes = lay.observacoes

    linhas = itertools.chain(([] if lay.tem_cabecalho else [primeira]), consumidas, it)

    if not lay.ok:
        # Não dá pra ler — mas CONTA as linhas, pra que o veto enxergue a perda.
        conta.lida = False
        conta.motivo = lay.motivo or 'layout não reconhecido'
        for linha in linhas:
            if any(linha):
                conta.lidas += 1
        return []

    m = lay.mapa
    i_uso = m.get('descricao_uso_iptu')
    i_area = m['area_construida_m2'] if 'area_construida_m2' in m else None
    registros = []
    for linha in linhas:
        if not any(linha):
            continue
        conta.lidas += 1
        if i_uso is not None:
            uso_bruto = str(_cel(linha, i_uso) or '').upper()
            if any(k in uso_bruto for k in ab.USOS_RESIDENCIAIS_KW):
                conta.residenciais_lidas += 1

        vals = {}
        for campo in CAMPOS_ORIGEM:
            idx = m.get(campo)
            bruto = _cel(linha, idx)
            tipo = TIPO_CAMPO[campo]
            if tipo == 'num':
                vals[campo] = ab.parse_float(bruto)
            elif tipo == 'data':
                vals[campo] = ab.parse_data(bruto)
            elif tipo == 'numero':
                vals[campo] = ab.normalize_numero(ab.parse_str(bruto))
            else:
                vals[campo] = ab.parse_str(bruto)

        # Só não entra a linha da qual nem o valor se consegue ler — aí o layout falhou de
        # verdade. Cadastro (SQL) vazio NÃO descarta mais: a linha entra assim mesmo.
        # (Plano técnico §3.3(h): essas linhas TAMBÉM entram.)
        if vals['valor_transacao'] is None:
            conta.descarte('linha sem valor legível — não entendida')
            continue
        conta.entendidas += 1
        if not vals['sql']:
            conta.sem_cadastro += 1

        area = vals['area_construida_m2']
        valor = vals['valor_transacao']
        # Salvaguarda do plano: abaixo de R$100 a linha entra, mas SEM valor por m² —
        # senão um preço simbólico envenena qualquer média.
        if area and area > 0 and valor >= VALOR_MINIMO:
            vals['valor_m2'] = round(valor / area, 2)
        else:
            vals['valor_m2'] = None
        if valor < VALOR_MINIMO:
            conta.abaixo_do_piso += 1
        data = vals['data_transacao']
        try:
            vals['ano'] = int(data[:4]) if data else None
        except (ValueError, TypeError):
            vals['ano'] = None
        norm = normalize_logradouro_v2(vals['logradouro'])
        vals['logradouro_norm'] = norm
        vals['logradouro_fmt'] = nb.title_case_logradouro(norm) if norm else None
        vals['cartorio'] = ab.normalize_cartorio(vals['cartorio'])
        uso = vals['descricao_uso_iptu'] or ''
        vals['residencial'] = 1 if any(k in uso for k in ab.USOS_RESIDENCIAIS_KW) else 0
        vals['ano_planilha'] = ano_planilha
        vals['aba_origem'] = nome_aba
        # Impressão digital da declaração. A MATRÍCULA entra na chave desde 06/08/2026
        # (decisão do Alex), e é a única diferença em relação à chave usada até aqui.
        #
        # Por quê: sem ela, duas vagas de garagem do mesmo prédio, vendidas no mesmo dia,
        # pelo mesmo valor e com o complemento escrito igual viravam UM registro só —
        # eram imóveis diferentes, e um sumia. Medido nas 21 planilhas: **7.602 vendas**
        # perdidas assim (0,29% da base), concentradas justamente em vaga e em unidade de
        # lançamento, que é onde o ACM mais erra. A matrícula é o registro do imóvel no
        # cartório: matrícula diferente = imóvel diferente.
        #
        # O risco que isso abre, e que é menor de propósito: se a Prefeitura republicar
        # uma declaração com a matrícula corrigida, ela entra duas vezes. Repetição a
        # gente enxerga e conserta; o que some, ninguém procura.
        #
        # ATENÇÃO: `atualizar_banco.py` (a rotina mensal) PRECISA usar exatamente esta
        # mesma chave. Se as duas divergirem, a rodada do dia 10 reinsere como novidade
        # o que já está na base. Foi assim que 124 mil revendas se perderam em 2026.
        vals['chave_dedup'] = '\x1f'.join([
            vals['sql'] or '', vals['numero'] or '', vals['complemento'] or '',
            (data or '')[:19], f'{valor:.2f}', vals['matricula'] or '',
        ])
        registros.append(vals)
    return registros


# ─── Banco ──────────────────────────────────────────────────────────────────────
COLS_BANCO = CAMPOS_ORIGEM + ['valor_m2', 'ano', 'logradouro_norm', 'logradouro_fmt',
                              'residencial', 'ano_planilha', 'aba_origem', 'chave_dedup']

TIPO_SQLITE = {'num': 'REAL', 'data': 'TEXT', 'txt': 'TEXT', 'numero': 'TEXT'}


def criar_banco(caminho):
    if os.path.exists(caminho):
        os.remove(caminho)
    conn = sqlite3.connect(caminho)
    # Página maior deixa o índice de deduplicação mais raso (menos leituras por gravação) e não
    # custa memória. O cache fica DELIBERADAMENTE modesto: nesta máquina a memória vive no
    # limite, e pedir cache grande faz o sistema paginar em disco — que é justamente o que se
    # quer evitar. Ver a nota de velocidade no relatório.
    conn.execute('PRAGMA page_size=8192')
    conn.execute('PRAGMA journal_mode=OFF')
    conn.execute('PRAGMA synchronous=OFF')
    conn.execute('PRAGMA cache_size=-48000')    # ~48 MB, sem forçar paginação
    defs = [f'{c} {TIPO_SQLITE[TIPO_CAMPO[c]]}' for c in CAMPOS_ORIGEM]
    defs += ['valor_m2 REAL', 'ano INTEGER', 'logradouro_norm TEXT', 'logradouro_fmt TEXT',
             'residencial INTEGER', 'ano_planilha INTEGER', 'aba_origem TEXT',
             'chave_dedup TEXT']
    conn.execute('CREATE TABLE vendas (id INTEGER PRIMARY KEY, ' + ', '.join(defs) + ')')
    # Deduplicação no arquivo, não na memória: (sql, numero, complemento, data, valor).
    conn.execute('CREATE UNIQUE INDEX idx_dedup ON vendas(chave_dedup)')
    conn.commit()
    return conn


INSERT_SQL = ('INSERT OR IGNORE INTO vendas (' + ','.join(COLS_BANCO) + ') VALUES ('
              + ','.join(['?'] * len(COLS_BANCO)) + ')')


def gravar(conn, registros):
    """Grava em bloco e devolve (novos, repetidos)."""
    if not registros:
        return 0, 0
    cur = conn.cursor()
    antes = conn.total_changes
    cur.executemany(INSERT_SQL, [[r.get(c) for c in COLS_BANCO] for r in registros])
    conn.commit()
    novos = conn.total_changes - antes   # INSERT OR IGNORE: repetida não conta como mudança
    return novos, len(registros) - novos


def indexar(conn):
    for col in ['ano', 'bairro', 'logradouro_norm', 'referencia', 'sql', 'residencial',
                'ano_planilha']:
        conn.execute(f'CREATE INDEX idx_{col} ON vendas({col})')
    conn.commit()


# ─── Boletim (lição L-21) ───────────────────────────────────────────────────────
def reais(v):
    """Formata dinheiro do jeito brasileiro: 1.234.567,89 (o Alex lê este boletim)."""
    if v is None:
        return '—'
    return f'{v:,.2f}'.translate(str.maketrans({',': '.', '.': ','}))


def escrever_boletim(caminho, saida_db, planilhas, segundos, falhas, conn=None):
    L = []
    L.append('# Boletim de importação — ITBI v2')
    L.append('')
    L.append(f'> Gerado em {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} · '
             f'banco: `{saida_db}` · tempo: {segundos:.0f}s')
    L.append('')
    tot = {k: 0 for k in ('lidas', 'entendidas', 'novas', 'repetidas', 'residenciais_lidas',
                          'abaixo_do_piso', 'sem_cadastro')}
    for _, _, contas in planilhas:
        for c in contas:
            for k in tot:
                tot[k] += getattr(c, k)
    L.append('## 1. Resumo')
    L.append('')
    L.append('| medida | valor |')
    L.append('|---|---:|')
    L.append(f'| planilhas lidas | {len(planilhas)} |')
    L.append(f'| linhas lidas | {tot["lidas"]:,} |')
    L.append(f'| linhas entendidas | {tot["entendidas"]:,} |')
    L.append(f'| linhas aproveitadas (novas no banco) | {tot["novas"]:,} |')
    L.append(f'| linhas repetidas (já estavam) | {tot["repetidas"]:,} |')
    L.append(f'| linhas residenciais entre as lidas | {tot["residenciais_lidas"]:,} |')
    L.append(f'| — dessas, com valor abaixo de R$ {reais(VALOR_MINIMO)} (entraram, sem valor '
             f'por m²) | {tot["abaixo_do_piso"]:,} |')
    L.append(f'| — dessas, sem cadastro (SQL) na origem (entraram assim mesmo) | '
             f'{tot["sem_cadastro"]:,} |')
    L.append('')
    L.append(f'**A conta fecha assim:** {tot["lidas"]:,} lidas − {tot["repetidas"]:,} repetidas '
             f'− {sum(sum(c.descartes.values()) for _, _, cs in planilhas for c in cs):,} '
             f'não entendidas = {tot["novas"]:,} no banco.')
    L.append('')

    # ── Uma linha por planilha: é o que se confere contra o gabarito medido por fora ──
    L.append('## 2. Uma linha por planilha (para conferir contra o gabarito)')
    L.append('')
    L.append('| planilha | linhas lidas | residenciais lidas | entendidas | novas | repetidas |')
    L.append('|---|---:|---:|---:|---:|---:|')
    for nome, _, contas in planilhas:
        d = [c for c in contas if c.eh_dados]
        L.append(f'| {nome} | {sum(c.lidas for c in d):,} | {sum(c.residenciais_lidas for c in d):,} '
                 f'| {sum(c.entendidas for c in d):,} | {sum(c.novas for c in d):,} '
                 f'| {sum(c.repetidas for c in d):,} |')
    L.append(f'| **TOTAL** | **{tot["lidas"]:,}** | **{tot["residenciais_lidas"]:,}** | '
             f'**{tot["entendidas"]:,}** | **{tot["novas"]:,}** | **{tot["repetidas"]:,}** |')
    L.append('')

    # ── Detector de buraco por mês ──────────────────────────────────────────────
    L.append('## 3. Meses fora da curva (buraco possível)')
    L.append('')
    L.append(f'Regra: mês com menos de {FRACAO_MES_SUSPEITO:.0%} da mediana de linhas entendidas '
             'do próprio ano. É o detector que teria pego JAN/OUT-2024 há um ano — nem todo '
             'aviso é defeito, mas todo aviso precisa de explicação.')
    L.append('')
    avisos = []
    for nome, _, contas in planilhas:
        d = [c for c in contas if c.eh_dados]
        if len(d) < 3:
            continue
        med = statistics.median([c.entendidas for c in d])
        corte = med * FRACAO_MES_SUSPEITO
        for c in d:
            if c.entendidas < corte:
                avisos.append((nome, c.nome, c.entendidas, med, corte))
    if avisos:
        L.append('| planilha | aba | entendidas | mediana do ano | corte | quanto da mediana |')
        L.append('|---|---|---:|---:|---:|---:|')
        for nome, aba, ent, med, corte in avisos:
            pct = (ent / med * 100) if med else 0
            L.append(f'| {nome} | {aba} | {ent:,} | {med:,.0f} | {corte:,.0f} | {pct:.0f}% |')
    else:
        L.append('- nenhum mês abaixo do corte.')
    L.append('')

    # ── Valores impossíveis: não filtra, sinaliza ───────────────────────────────
    L.append(f'## 4. Valores acima de R$ {reais(LIMIAR_VALOR_ABSURDO)} (sinalizados, NÃO descartados)')
    L.append('')
    L.append('Decisão do Alex: guardar tudo e poder ver tudo. Um valor impossível é uma '
             'declaração real de um contribuinte real — pode ser erro de digitação dele, e quem '
             'decide o que fazer é o Alex. Aqui só contamos e mostramos.')
    L.append('')
    if conn is not None:
        n = conn.execute('SELECT COUNT(*) FROM vendas WHERE valor_transacao > ?',
                         (LIMIAR_VALOR_ABSURDO,)).fetchone()[0]
        L.append(f'- registros acima do limiar: **{n:,}**')
        L.append('')
        if True:   # os dez maiores saem SEMPRE, mesmo que nenhum passe do limiar —
            #        é o que permite conferir se o teto está no lugar certo
            L.append('**Os dez maiores valores da base (acima ou não do limiar):**')
            L.append('')
            L.append('| valor declarado | logradouro | nº | compl. | bairro | data | uso | origem |')
            L.append('|---:|---|---|---|---|---|---|---|')
            for r in conn.execute(
                    'SELECT valor_transacao, logradouro_fmt, numero, complemento, bairro, '
                    'data_transacao, descricao_uso_iptu, ano_planilha, aba_origem '
                    'FROM vendas ORDER BY valor_transacao DESC LIMIT 10'):
                L.append(f'| {reais(r[0])} | {r[1] or "—"} | {r[2] or "—"} | {(r[3] or "—")[:20]} | '
                         f'{(r[4] or "—")[:20]} | {(r[5] or "—")[:10]} | {(r[6] or "—")[:28]} | '
                         f'{r[7]}/{r[8]} |')
            L.append('')

        # ── Alarme relativo: os dez maiores por valor de metro quadrado ─────────
        # Só no recorte residencial: fora dele (terreno, galpão) o preço por metro não tem
        # padrão nenhum e a lista viraria ruído. Só lista — nunca descarta.
        L.append('## 5. Os dez maiores por valor do metro quadrado (só residencial)')
        L.append('')
        L.append('Alarme relativo, ao lado do limiar absoluto da seção 4: um erro de digitação '
                 'do contribuinte costuma aparecer aqui antes de aparecer lá. **Só listamos, '
                 'nada é descartado.**')
        L.append('')
        L.append('| valor por m² | valor declarado | área (m²) | logradouro | nº | compl. | '
                 'bairro | data | origem |')
        L.append('|---:|---:|---:|---|---|---|---|---|---|')
        for r in conn.execute(
                'SELECT valor_m2, valor_transacao, area_construida_m2, logradouro_fmt, numero, '
                'complemento, bairro, data_transacao, ano_planilha, aba_origem '
                'FROM vendas WHERE residencial = 1 AND valor_m2 IS NOT NULL '
                'ORDER BY valor_m2 DESC LIMIT 10'):
            L.append(f'| {reais(r[0])} | {reais(r[1])} | {reais(r[2])} | {r[3] or "—"} | '
                     f'{r[4] or "—"} | {(r[5] or "—")[:18]} | {(r[6] or "—")[:18]} | '
                     f'{(r[7] or "—")[:10]} | {r[8]}/{r[9]} |')
        L.append('')

        # ── Contagem do piso, por planilha ──────────────────────────────────────
        L.append(f'## 6. Linhas com valor abaixo de R$ {reais(VALOR_MINIMO)}')
        L.append('')
        L.append('Pelo plano técnico §3.3(h), **essas linhas entram no banco** — com duas '
                 'salvaguardas: `valor_m2` fica **vazio** (para não envenenar média nenhuma) e a '
                 'contagem aparece aqui. Nada é descartado por ser barato.')
        L.append('')
        L.append('| planilha | linhas abaixo do piso | sem cadastro (SQL) na origem |')
        L.append('|---|---:|---:|')
        for nome, _, contas in planilhas:
            d = [c for c in contas if c.eh_dados]
            L.append(f'| {nome} | {sum(c.abaixo_do_piso for c in d):,} | '
                     f'{sum(c.sem_cadastro for c in d):,} |')
        L.append(f'| **TOTAL** | **{tot["abaixo_do_piso"]:,}** | **{tot["sem_cadastro"]:,}** |')
        L.append('')
        n_m2 = conn.execute('SELECT COUNT(*) FROM vendas WHERE valor_transacao < ? '
                            'AND valor_m2 IS NOT NULL', (VALOR_MINIMO,)).fetchone()[0]
        L.append(f'Conferência da salvaguarda: registros abaixo do piso que ficaram COM valor '
                 f'por m² (tem de ser zero): **{n_m2}**')
        L.append('')

    for nome, ano, contas in planilhas:
        dados = [c for c in contas if c.eh_dados]
        nao_dados = [c for c in contas if not c.eh_dados]
        L.append(f'## Planilha {nome}')
        L.append('')
        L.append(f'- abas encontradas: {len(contas)}')
        L.append(f'- abas de dados: {len(dados)} · abas que não são de dados: '
                 f'{len(nao_dados)} ({", ".join(c.nome for c in nao_dados) or "—"})')
        L.append('')
        L.append('| aba | layout reconhecido | lidas | entendidas | novas | repetidas |')
        L.append('|---|---|---:|---:|---:|---:|')
        for c in dados:
            L.append(f'| {c.nome} | {c.layout} | {c.lidas:,} | {c.entendidas:,} | '
                     f'{c.novas:,} | {c.repetidas:,} |')
        L.append('')
        obs = [c for c in dados if c.observacoes]
        if obs:
            L.append('**Observações de layout (o que foi inferido ou corrigido pelo conteúdo):**')
            L.append('')
            for c in obs:
                for o in c.observacoes:
                    L.append(f'- `{c.nome}` — {o}')
            L.append('')
        L.append('**Abas de dados que não renderam nada (com o motivo):**')
        L.append('')
        problemas = [c for c in dados if c.situacao() in ('nao_lida', 'nada_entendido', 'vazia')]
        if problemas:
            for c in problemas:
                s = c.situacao()
                if s == 'nao_lida':
                    L.append(f'- `{c.nome}` — NÃO LIDA: {c.motivo} (**{c.lidas:,} linhas ficaram de fora**)')
                elif s == 'nada_entendido':
                    L.append(f'- `{c.nome}` — LIDA MAS SEM NADA ENTENDIDO: {c.lidas:,} linhas, '
                             f'0 entendidas (**falha — ver veto**)')
                else:
                    L.append(f'- `{c.nome}` — vazia na origem (mês sem lançamento). Não é falha.')
        else:
            L.append('- nenhuma: todas as abas de dados foram lidas e renderam registros.')
        L.append('')
        desc = {}
        for c in dados:
            for k, v in c.descartes.items():
                desc[k] = desc.get(k, 0) + v
        L.append('**Descartes por motivo:**')
        L.append('')
        if desc:
            L.append('| motivo | linhas |')
            L.append('|---|---:|')
            for k, v in sorted(desc.items(), key=lambda x: -x[1]):
                L.append(f'| {k} | {v:,} |')
        else:
            L.append('- nenhum.')
        L.append('')

    L.append('## Veto — aba com linhas e nada aproveitado')
    L.append('')
    if falhas:
        L.append('**FALHA.** As abas abaixo têm linhas e nenhuma foi entendida:')
        L.append('')
        for f in falhas:
            L.append(f'- {f}')
    else:
        L.append('OK — nenhuma aba com linhas ficou sem nada entendido.')
    L.append('')
    with open(caminho, 'w') as fh:
        fh.write('\n'.join(L))


# ─── Programa ───────────────────────────────────────────────────────────────────
def processar_planilha(caminho, ano_planilha, conn):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    contas = []
    try:
        for sn in wb.sheetnames:
            conta = ContaAba(sn)
            conta.eh_dados = eh_aba_de_dados(sn)
            contas.append(conta)
            if not conta.eh_dados:
                conta.lida = False
                conta.motivo = 'aba de apoio (não é de dados)'
                conta.layout = '—'
                continue
            regs = ler_aba(wb[sn], sn, ano_planilha, conta)
            novos, repetidos = gravar(conn, regs)
            conta.novas, conta.repetidas = novos, repetidos
            print(f'    {sn}: {conta.lidas:,} lidas | {conta.entendidas:,} entendidas | '
                  f'{novos:,} novas | {repetidos:,} repetidas | {conta.layout}', flush=True)
    finally:
        wb.close()
    return contas


def expandir_anos(tokens):
    anos = []
    for t in tokens:
        if '-' in t:
            a, b = t.split('-', 1)
            anos.extend(range(int(a), int(b) + 1))
        else:
            anos.append(int(t))
    return anos


def main():
    p = argparse.ArgumentParser(description='Leitor v2 das planilhas do ITBI (todos os tipos).')
    p.add_argument('--anos', nargs='+', help='ex.: 2019 2024  ou  2019-2026')
    p.add_argument('--arquivos', nargs='+', help='caminhos de .xlsx (em vez de --anos)')
    p.add_argument('--saida', required=True, help='arquivo .db a gerar')
    p.add_argument('--boletim', help='arquivo .md do boletim (padrão: <saida>_boletim.md)')
    args = p.parse_args()

    if not args.anos and not args.arquivos:
        p.error('informe --anos ou --arquivos')

    alvos = []
    if args.anos:
        for ano in expandir_anos(args.anos):
            alvos.append((os.path.join(PLAN_DIR, f'{ano}.xlsx'), ano))
    for caminho in (args.arquivos or []):
        base = os.path.basename(caminho)
        m = re.match(r'(\d{4})', base)
        alvos.append((caminho, int(m.group(1)) if m else None))

    boletim = args.boletim or (os.path.splitext(args.saida)[0] + '_boletim.md')
    t0 = time.time()
    conn = criar_banco(args.saida)
    planilhas = []
    for caminho, ano in alvos:
        if not os.path.exists(caminho):
            print(f'  (sem {caminho})')
            continue
        print(f'  {os.path.basename(caminho)}:', flush=True)
        t1 = time.time()
        contas = processar_planilha(caminho, ano, conn)
        planilhas.append((os.path.basename(caminho), ano, contas))
        print(f'    -> {time.time() - t1:.0f}s', flush=True)
    indexar(conn)

    falhas = []
    for nome, _, contas in planilhas:
        for c in contas:
            if c.eh_dados and c.lidas > 0 and c.entendidas == 0:
                falhas.append(f'`{nome}` / `{c.nome}`: {c.lidas:,} linhas lidas, '
                              f'0 entendidas — {c.motivo or c.layout}')
    segundos = time.time() - t0
    escrever_boletim(boletim, args.saida, planilhas, segundos, falhas, conn)

    total = conn.execute('SELECT COUNT(*) FROM vendas').fetchone()[0]
    conn.close()
    print(f'\nBanco: {args.saida} ({total:,} registros, '
          f'{os.path.getsize(args.saida) / 1024 / 1024:.0f} MB)')
    print(f'Boletim: {boletim}')
    print(f'Tempo: {segundos:.0f}s')
    if falhas:
        print('\nERRO — aba com linhas e nada entendido:')
        for f in falhas:
            print('  - ' + f.replace('`', ''))
        sys.exit(1)
    print('OK — nenhuma aba com linhas ficou sem nada entendido.')


if __name__ == '__main__':
    main()
