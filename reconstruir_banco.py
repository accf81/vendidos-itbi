#!/usr/bin/env python3
"""
Reconstrução do banco ITBI a partir das planilhas oficiais (processo único).
=========================================================================
Gera um banco NOVO, limpo e consistente, aplicando todas as regras decididas em
Dados ITBI/PROCESSO_RECONSTRUCAO.md. NÃO toca no banco de produção: escreve num
arquivo de saída à parte, pra ser validado antes de qualquer troca.

Regras aplicadas (reusa as funções já existentes — nada reescrito à mão):
  - filtro residencial (inclui LOJA E RESIDÊNCIA, decisão do Alex)   [ab]
  - valor com correção do bug x100; descarta < R$100                 [ab.parse_float]
  - número 99999 -> S/N                                              [aqui]
  - normalização de logradouro + 34 correções                       [ab.normalize_logradouro]
  - cartório -> "N CRI"                                              [ab.normalize_cartorio]
  - capitalização uniforme (MAIÚSCULAS)                              [ab.parse_str]
  - ano derivado da data                                            [aqui]
  - deduplicação nova: (sql, numero, comp, data, valor)             [aqui]
  - valor_m2 = valor/area (sempre calculado)                        [aqui]
  - logradouro_fmt (versão de exibição)                             [nb.title_case_logradouro]

Uso:
  python3 reconstruir_banco.py 2019 2026 /caminho/saida.db
"""
import sqlite3, os, sys, time, csv
import openpyxl
import atualizar_banco as ab
import normalizar_banco as nb

PLAN_DIR = "/Users/accf81/Documents/IA/Claude/Projects/Dados ITBI/planilhas_originais"
AREA_MIN, AREA_MAX = 8, 2000  # faixa plausível pra apartamento/casa (fora disso: guardar p/ investigar)

# Ordem EXATA das colunas do banco de produção (drop-in replacement)
COLS = ['sql','data_transacao','logradouro','numero','complemento','bairro','referencia',
        'area_construida_m2','valor_transacao','valor_m2','cartorio','matricula','ano',
        'logradouro_norm','logradouro_fmt','proporcao_transmitida']
COL_PROP_NOME = 'Proporção Transmitida (%)'  # % do imóvel transmitida na transação (100 = imóvel inteiro)

norm_num = ab.normalize_numero  # '99999' -> S/N e tira o '.0' do número que vem como
                                # número do Excel (planilha de 2024) — ver ab.normalize_numero

def processar(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out=[]
    for sn in wb.sheetnames:
        ws=wb[sn]; it=ws.iter_rows(values_only=True); h=next(it,None)
        if not h: continue
        ci={}
        for i,nm in enumerate(h):
            if nm is None: continue
            k=str(nm).strip()
            if k in ab.COLUNAS_XLSX: ci[ab.COLUNAS_XLSX[k]]=i
            if k == COL_PROP_NOME: ci['proporcao']=i
        if not all(x in ci for x in ['sql','logradouro','valor_transacao','data_transacao','_uso']):
            continue
        for row in it:
            if not any(row): continue
            uso=ab.parse_str(row[ci['_uso']]) or ''
            if not any(kw in uso for kw in ab.USOS_RESIDENCIAIS_KW): continue
            sql=ab.parse_str(row[ci['sql']])
            if not sql: continue
            def g(f,p=ab.parse_str):
                idx=ci.get(f); return p(row[idx]) if idx is not None else None
            numero=norm_num(g('numero')); comp=g('complemento')
            valor=g('valor_transacao', ab.parse_float)
            if valor is None or valor<100: continue
            data=g('data_transacao', ab.parse_data)
            logr=g('logradouro'); bairro=g('bairro'); ref=g('referencia')
            cart=g('cartorio'); matr=g('matricula'); area=g('area_construida_m2', ab.parse_float)
            valor_m2=round(valor/area,2) if area and area>0 else None
            try: ano=int((data or '')[:4]) if data else None
            except: ano=None
            logr_norm=ab.normalize_logradouro(logr)
            logr_fmt=nb.title_case_logradouro(logr_norm) if logr_norm else None
            prop=g('proporcao', ab.parse_float)
            registro=(sql,data,logr,numero,comp,bairro,ref,area,valor,valor_m2,
                      ab.normalize_cartorio(cart),matr,ano,logr_norm,logr_fmt,prop)
            out.append(registro)
    return out

def main():
    if len(sys.argv)<4:
        print("Uso: python3 reconstruir_banco.py <ano_ini> <ano_fim> <saida.db>"); sys.exit(1)
    ai, af, saida = int(sys.argv[1]), int(sys.argv[2]), sys.argv[3]
    t0=time.time()
    vistos={}  # chave nova -> registro
    for ano in range(ai, af+1):
        p=os.path.join(PLAN_DIR,f"{ano}.xlsx")
        if not os.path.exists(p): print(f"  (sem {ano})"); continue
        regs=processar(p)
        for r in regs:
            chave=(r[0], r[3], r[4], (r[1] or '')[:19], r[8])  # sql,numero,comp,data,valor
            if chave not in vistos: vistos[chave]=r
        print(f"  {ano}: {len(regs):,} lidos | únicos: {len(vistos):,} ({time.time()-t0:.0f}s)", flush=True)

    registros=list(vistos.values())
    # Gravar banco novo (schema idêntico ao de produção)
    if os.path.exists(saida): os.remove(saida)
    conn=sqlite3.connect(saida); cur=conn.cursor()
    cur.execute("""CREATE TABLE vendas (
        sql TEXT, data_transacao TEXT, logradouro TEXT, numero TEXT,
        complemento TEXT, bairro TEXT, referencia TEXT,
        area_construida_m2 REAL, valor_transacao REAL, valor_m2 REAL,
        cartorio TEXT, matricula TEXT, ano INTEGER
    , logradouro_norm TEXT, logradouro_fmt TEXT, proporcao_transmitida REAL)""")
    cur.executemany(f"INSERT INTO vendas ({','.join(COLS)}) VALUES ({','.join(['?']*len(COLS))})", registros)
    for col in ['logradouro','referencia','bairro','ano','logradouro_norm']:
        cur.execute(f"CREATE INDEX idx_{col} ON vendas({col})")
    conn.commit()

    # Relatório de áreas suspeitas (guardar p/ investigar — decisão do Alex: mantém o preço)
    rel = saida.replace('.db','') + '_areas_suspeitas.csv'
    peq=gra=0
    with open(rel,'w',newline='') as f:
        w=csv.writer(f); w.writerow(['motivo','sql','numero','complemento','logradouro_norm','area_m2','valor','valor_m2','ano'])
        for r in registros:
            area=r[7]
            if area is None: continue
            if area < AREA_MIN:
                w.writerow(['area_menor_que_8', r[0],r[3],r[4],r[13],area,r[8],r[9],r[12]]); peq+=1
            elif area > AREA_MAX:
                w.writerow(['area_maior_que_2000', r[0],r[3],r[4],r[13],area,r[8],r[9],r[12]]); gra+=1
    conn.close()
    print(f"\n✅ Banco novo: {saida}  ({len(registros):,} registros, {os.path.getsize(saida)/1024/1024:.0f} MB)")
    print(f"   Áreas suspeitas guardadas em {os.path.basename(rel)}: {peq:,} muito pequenas (<8m²), {gra:,} muito grandes (>2000m²)")
    print(f"   Tempo: {time.time()-t0:.0f}s")

if __name__=='__main__':
    main()
