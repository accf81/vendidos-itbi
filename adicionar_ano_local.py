#!/usr/bin/env python3
"""
Adicionar ano histórico ao banco ITBI — a partir de planilha LOCAL já arquivada.
=============================================================================
Reusa 100% das REGRAS de atualizar_banco.py (importadas: filtro residencial,
normalização de logradouro/cartório, tratamento do valor com a correção do bug
x100, deduplicação por (sql, numero, complemento) contra o banco inteiro, e a
sincronização com o Supabase). A ÚNICA diferença é a origem do dado: em vez de
baixar o XLSX do ano corrente da Prefeitura, lê um arquivo já baixado e guardado
em Dados ITBI/planilhas_originais/AAAA.xlsx.

Feito pra carregar anos históricos (2019, 2018, ...) sem depender do site da
Prefeitura e sem tocar no script de produção mensal.

Uso (rodar de dentro da pasta Pandora Data SP, onde vive o banco):
  python3 adicionar_ano_local.py "/caminho/para/2019.xlsx"
  python3 adicionar_ano_local.py "/caminho/para/2019.xlsx" --dry-run   # só relata, não grava
"""
import sqlite3, gzip, shutil, os, sys, time
from datetime import datetime
import openpyxl
import atualizar_banco as ab

DB_FILE = ab.DB_FILE
TMP_DB  = '/tmp/itbi_add_ano.db'


def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    dry_run = '--dry-run' in sys.argv
    if not args:
        print("Uso: python3 adicionar_ano_local.py \"/caminho/AAAA.xlsx\" [--dry-run]")
        sys.exit(1)
    xlsx_path = args[0]
    if not os.path.exists(xlsx_path):
        print(f"✗ Arquivo não encontrado: {xlsx_path}"); sys.exit(1)

    backup_file = f'ITBI_SP_residencial_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db.gz'

    print(f"\n{'='*60}")
    print(f"  CARGA DE ANO HISTÓRICO — {os.path.basename(xlsx_path)}")
    if dry_run: print(f"  (DRY-RUN: só relata, não grava nada)")
    print(f"{'='*60}\n")

    # 1. Backup e descomprimir banco atual
    if not dry_run:
        print(f"1. Backup → {backup_file}")
        shutil.copy2(DB_FILE, backup_file)
        print(f"   ✓ {os.path.getsize(backup_file)/1024/1024:.1f} MB")

    print(f"\n2. Descomprimindo banco...")
    with gzip.open(DB_FILE, 'rb') as f: data = f.read()
    with open(TMP_DB, 'wb') as f: f.write(data)
    conn = sqlite3.connect(TMP_DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM vendas"); total_antes = cur.fetchone()[0]
    print(f"   Registros antes: {total_antes:,}")

    # 2. Chaves existentes p/ deduplicação (mesma chave do script mensal)
    print(f"\n3. Carregando chaves existentes...")
    cur.execute("SELECT sql, numero, complemento FROM vendas WHERE sql IS NOT NULL")
    chaves_existentes = {(r[0], r[1], r[2]) for r in cur.fetchall()}
    print(f"   ✓ {len(chaves_existentes):,} registros carregados")

    # 3. Ler o XLSX local (todas as abas mensais)
    print(f"\n4. Lendo XLSX local: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"   Abas encontradas: {wb.sheetnames}")

    novos = []
    usos_encontrados = set()
    anos_inseridos = {}
    total_lido = total_residencial = total_duplicata = total_invalido = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue

        col_idx = {}
        for i, nome in enumerate(header):
            if nome is None: continue
            nome_str = str(nome).strip()
            if nome_str in ab.COLUNAS_XLSX:
                col_idx[ab.COLUNAS_XLSX[nome_str]] = i

        obrigatorias = ['sql', 'logradouro', 'valor_transacao', 'data_transacao', '_uso']
        faltando = [c for c in obrigatorias if c not in col_idx]
        if faltando:
            print(f"   ⚠ Aba '{sheet_name}' ignorada (não é aba de dados): faltam {faltando}")
            continue

        for row in rows_iter:
            if not any(row): continue
            total_lido += 1

            uso = ab.parse_str(row[col_idx['_uso']]) or ''
            usos_encontrados.add(uso)
            if not any(kw in uso for kw in ab.USOS_RESIDENCIAIS_KW):
                continue
            total_residencial += 1

            sql_val = ab.parse_str(row[col_idx['sql']])
            if not sql_val:
                total_invalido += 1; continue

            def get(field, parser=ab.parse_str):
                idx = col_idx.get(field)
                return parser(row[idx]) if idx is not None else None

            logradouro  = get('logradouro')
            numero      = get('numero')
            complemento = get('complemento')
            numero_norm = 'S/N' if str(numero or '').strip() == '99999' else str(numero or '').strip()

            chave = (sql_val, numero_norm, complemento)
            if chave in chaves_existentes:
                total_duplicata += 1; continue

            bairro     = get('bairro')
            referencia = get('referencia')
            valor      = get('valor_transacao', ab.parse_float)
            data_str   = get('data_transacao', ab.parse_data)
            cartorio   = get('cartorio')
            matricula  = get('matricula')
            area       = get('area_construida_m2', ab.parse_float)

            if valor is None or valor < 100:
                total_invalido += 1; continue

            logradouro_norm = ab.normalize_logradouro(logradouro)
            cartorio_norm   = ab.normalize_cartorio(cartorio)
            valor_m2        = round(valor / area, 2) if area and area > 0 else None
            try:
                ano = int(data_str[:4]) if data_str else None
            except:
                ano = None
            anos_inseridos[ano] = anos_inseridos.get(ano, 0) + 1

            novos.append((
                sql_val, data_str, logradouro, numero_norm, complemento,
                bairro, referencia, area, valor, valor_m2,
                cartorio_norm, matricula, ano, logradouro_norm
            ))
            chaves_existentes.add(chave)

    print(f"\n   Estatísticas de leitura:")
    print(f"   Total linhas lidas:    {total_lido:,}")
    print(f"   Residenciais:          {total_residencial:,}")
    print(f"   Duplicatas (já no DB): {total_duplicata:,}")
    print(f"   Inválidos:             {total_invalido:,}")
    print(f"   → Novos a inserir:     {len(novos):,}")
    print(f"\n   Distribuição por ano dos novos: {dict(sorted(anos_inseridos.items(), key=lambda x:(x[0] is None,x[0])))}")

    if dry_run:
        print(f"\n   DRY-RUN — nada gravado. Fechando.")
        conn.close(); os.remove(TMP_DB)
        return

    if not novos:
        print("\n✅ Nenhum registro novo.")
        conn.close(); os.remove(TMP_DB); os.remove(backup_file)
        return

    print(f"\n5. Inserindo {len(novos):,} registros no banco local...")
    cur.executemany("""
        INSERT OR IGNORE INTO vendas
          (sql, data_transacao, logradouro, numero, complemento,
           bairro, referencia, area_construida_m2, valor_transacao, valor_m2,
           cartorio, matricula, ano, logradouro_norm)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, novos)
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM vendas"); total_depois = cur.fetchone()[0]
    inseridos = total_depois - total_antes
    print(f"   ✓ {inseridos:,} inseridos (total: {total_depois:,})")
    conn.close()

    ab.sincronizar_supabase(novos)

    print(f"\n6. Recomprimindo banco...")
    t0 = time.time()
    with open(TMP_DB, 'rb') as f_in:
        with gzip.open(DB_FILE, 'wb', compresslevel=6) as f_out:
            shutil.copyfileobj(f_in, f_out)
    print(f"   ✓ {os.path.getsize(DB_FILE)/1024/1024:.1f} MB em {time.time()-t0:.1f}s")
    os.remove(TMP_DB)

    print(f"\n{'─'*45}")
    print(f"  Registros antes:  {total_antes:,}")
    print(f"  Novos inseridos:  {inseridos:,}")
    print(f"  Total final:      {total_depois:,}")
    print(f"  Backup mantido:   {backup_file}")
    print(f"{'─'*45}")
    print(f"\n✅ CONCLUÍDO! Próximo passo: publicar o site (deploy).\n")


if __name__ == '__main__':
    main()
